from copy import copy
from itertools import filterfalse

from openpyxl import load_workbook
from openpyxl.styles import Font, colors

from excel_match_pattern import (
    many_names_years_pattern,
    name_pattern,
    names_pattern,
    names_years_pattern,
    year_pattern,
    years_pattern,
)
from word import save, search_and_paint


def parse_authors_and_years(authors_years_group):
    for authors_years in authors_years_group:
        authors_and_years = map(str.strip, authors_years.split(","))
        authors = next(authors_and_years).split()
        years = tuple(authors_and_years)
        yield tuple(filterfalse(lambda s: s in ("and", "et", "al."), authors)), years


def change_cell_font_color_to_red(row, col):
    cell = ws.cell(row_no, col_no)
    problematic_cell_Font = copy(cell.font)
    problematic_cell_Font.color = colors.RED
    cell.font = problematic_cell_Font


wb = load_workbook("table2.xlsx")
ws = wb.active
col_no = 17

for col in ws.iter_cols(
    min_col=col_no, max_col=col_no, min_row=5, max_row=221, values_only=True
):
    val: str
    for i, val in enumerate(col):
        row_no = i + 5
        if val:  # skip None
            if not many_names_years_pattern.fullmatch(val):
                print(row_no)
                raise ValueError(val)
            # print(row_no, end=": ")
            for names_years in names_years_pattern.findall(val):
                names = names_pattern.fullmatch(names_years[0])[1].split(" and ")
                years = year_pattern.findall(years_pattern.match(names_years[1])[1])
                for i, name in enumerate(names):
                    if " " in name:
                        first, last = name.split()
                        if str.isupper(first):
                            first, last = last, first
                        last = ".".join(filter(str.isalpha, last)) + "."
                        names[i] = first + ", " + last
                years_expanded = []
                for year in years:
                    if ", " in year:
                        years_expanded.extend(
                            year[:4] + c for c in year[4:].split(", ")
                        )
                    else:
                        years_expanded.append(year)
                years = years_expanded
                # print(names, years, end=", ")
                if not search_and_paint(names, years):
                    print(row_no)
                    # change_cell_font_color_to_red(row_no, col_no)
            # print()

save()
# wb.save("new_table2.xlsx")
